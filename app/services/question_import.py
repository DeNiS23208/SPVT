"""Импорт вопросов теста из Excel."""
from __future__ import annotations

import io
import json
from typing import Any

from openpyxl import Workbook, load_workbook
from sqlalchemy import func
from sqlalchemy.orm import Session

from app.models import Question, QuestionType, TestTicket
from app.services.test_tickets import default_ticket_for_test


def _parse_question_type(value: Any) -> QuestionType:
    text = str(value or "").strip().lower()
    if text in {"yes_no", "да/нет", "данет", "yesno"}:
        return QuestionType.yes_no
    return QuestionType.single_choice


def _parse_options(value: Any, question_type: QuestionType) -> list[str]:
    if question_type == QuestionType.yes_no:
        return ["Да", "Нет"]
    raw = str(value or "").strip()
    if not raw:
        return []
    if "|" in raw:
        return [part.strip() for part in raw.split("|") if part.strip()]
    if ";" in raw:
        return [part.strip() for part in raw.split(";") if part.strip()]
    return [raw]


def _row_to_question(row: tuple, sort_order: int) -> tuple[dict | None, str | None]:
    if not row or not str(row[0] or "").strip():
        return None, None

    text = str(row[0]).strip()
    qtype = _parse_question_type(row[1] if len(row) > 1 else None)
    options = _parse_options(row[2] if len(row) > 2 else None, qtype)
    correct = str(row[3] if len(row) > 3 else "").strip()
    # Старый шаблон: колонка 5 — «критический», колонка 6 — порядок; новый — порядок в колонке 5
    if len(row) > 5:
        order_raw = row[5]
    elif len(row) > 4:
        order_raw = row[4]
    else:
        order_raw = None
    try:
        order = int(order_raw) if order_raw not in (None, "") else sort_order
    except (TypeError, ValueError):
        order = sort_order

    if not correct:
        return None, f"«{text[:40]}…»: не указан правильный ответ"
    if qtype == QuestionType.single_choice and len(options) < 2:
        return None, f"«{text[:40]}…»: нужно минимум 2 варианта ответа"
    if correct not in options:
        return None, f"«{text[:40]}…»: правильный ответ должен совпадать с одним из вариантов"

    return {
        "text": text,
        "question_type": qtype,
        "options": options,
        "correct_answer": correct,
        "sort_order": order,
        "is_active": True,
    }, None


def import_questions_from_excel(db: Session, test_type_id: int, file_bytes: bytes) -> dict:
    wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(min_row=2, values_only=True))

    max_order = (
        db.query(func.max(Question.sort_order))
        .filter(Question.test_type_id == test_type_id)
        .scalar()
        or 0
    )

    created = 0
    errors: list[str] = []
    next_order = max_order + 1
    ticket = default_ticket_for_test(db, test_type_id)

    for idx, row in enumerate(rows, start=2):
        payload, err = _row_to_question(row, next_order)
        if err:
            errors.append(f"Строка {idx}: {err}")
            continue
        if not payload:
            continue

        db.add(
            Question(
                test_type_id=test_type_id,
                ticket_id=ticket.id,
                text=payload["text"],
                question_type=payload["question_type"],
                options_json=json.dumps(payload["options"], ensure_ascii=False),
                correct_answer=payload["correct_answer"],
                allow_multiple_correct=False,
                is_critical=False,
                sort_order=payload["sort_order"],
                is_active=True,
            )
        )
        created += 1
        next_order = max(next_order, payload["sort_order"]) + 1

    db.commit()
    return {"created": created, "errors": errors, "error_count": len(errors)}


def build_questions_template_xlsx() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Вопросы"
    ws.append(
        [
            "Вопрос",
            "Тип (yes_no / single_choice)",
            "Варианты (через | )",
            "Правильный ответ",
            "Порядок",
        ]
    )
    ws.append(
        [
            "Допускается ли работа при опьянении?",
            "yes_no",
            "Да|Нет",
            "Нет",
            1,
        ]
    )
    ws.append(
        [
            "Что сделать перед началом работы на высоте?",
            "single_choice",
            "Начать сразу|Пройти инструктаж и проверить СИЗ|Попросить коллегу",
            "Пройти инструктаж и проверить СИЗ",
            2,
        ]
    )
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_test_questions_export_xlsx(db: Session, test_type_id: int) -> bytes:
    """Выгрузка всех вопросов теста по билетам в Excel."""
    pairs = (
        db.query(Question, TestTicket)
        .outerjoin(TestTicket, Question.ticket_id == TestTicket.id)
        .filter(Question.test_type_id == test_type_id, Question.is_active.is_(True))
        .order_by(
            TestTicket.sort_order.nulls_last(),
            TestTicket.id,
            Question.sort_order,
            Question.id,
        )
        .all()
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Вопросы"
    ws.append(["Билет", "Вопрос", "Правильный ответ"])

    for question, ticket in pairs:
        ws.append(
            [
                ticket.title if ticket else "—",
                question.text,
                question.correct_answer,
            ]
        )

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
