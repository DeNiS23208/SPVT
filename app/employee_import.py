from __future__ import annotations

import os
from dataclasses import dataclass

from openpyxl import load_workbook
from sqlalchemy.orm import Session

from app.auth import hash_password
from app.models import User, UserRole
from app.name_utils import normalize_name_key
from app.usernames import DEFAULT_PASSWORD, username_from_name


@dataclass
class EmployeeRow:
    full_name: str
    position: str
    department: str


def read_employees_xlsx(path: str) -> list[EmployeeRow]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]
    rows = list(worksheet.iter_rows(values_only=True))
    if not rows:
        return []

    header = [str(cell or "").strip().lower() for cell in rows[0]]
    name_idx = next((i for i, value in enumerate(header) if "сотрудник" in value or value == "фио"), 0)
    position_idx = next((i for i, value in enumerate(header) if "должност" in value), 1)
    department_idx = next((i for i, value in enumerate(header) if "подраздел" in value), 2)

    employees: list[EmployeeRow] = []
    for row in rows[1:]:
        if not row or not row[name_idx]:
            continue
        employees.append(
            EmployeeRow(
                full_name=str(row[name_idx]).strip(),
                position=str(row[position_idx] or "").strip(),
                department=str(row[department_idx] or "").strip(),
            )
        )
    return employees


@dataclass
class ImportStats:
    created: int = 0
    updated: int = 0
    skipped: int = 0


def import_employees_from_xlsx(
    db: Session,
    path: str,
    *,
    default_password: str | None = None,
    update_existing: bool = True,
) -> ImportStats:
    password = default_password or os.environ.get("DEFAULT_WORKER_PASSWORD", DEFAULT_PASSWORD)
    password_hash = hash_password(password)
    employees = read_employees_xlsx(path)
    stats = ImportStats()

    existing_by_name = {
        normalize_name_key(user.full_name): user
        for user in db.query(User).filter(User.role == UserRole.worker).all()
    }
    used_usernames = {user.username for user in db.query(User).all()}

    for index, employee in enumerate(employees, start=1):
        key = normalize_name_key(employee.full_name)
        existing = existing_by_name.get(key)

        if existing:
            if existing.username in used_usernames:
                used_usernames.discard(existing.username)
            new_username = username_from_name(employee.full_name, index, used_usernames)
            if update_existing:
                changed = False
                if existing.username != new_username:
                    existing.username = new_username
                    changed = True
                if employee.position and existing.position != employee.position:
                    existing.position = employee.position
                    changed = True
                if employee.department and existing.department != employee.department:
                    existing.department = employee.department
                    changed = True
                existing.password_hash = password_hash
                if changed:
                    stats.updated += 1
                else:
                    stats.skipped += 1
            else:
                stats.skipped += 1
            continue

        new_username = username_from_name(employee.full_name, index, used_usernames)
        db.add(
            User(
                username=new_username,
                password_hash=password_hash,
                role=UserRole.worker,
                full_name=employee.full_name,
                position=employee.position or None,
                department=employee.department or None,
                is_active=True,
            )
        )
        stats.created += 1
        if stats.created % 250 == 0:
            db.flush()
            print(f"  импортировано {stats.created}...", flush=True)

    db.commit()
    return stats
